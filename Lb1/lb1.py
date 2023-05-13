import requests
from bs4 import BeautifulSoup
import openpyxl

url = 'https://omsk.mlsn.ru/pokupka-nedvizhimost/?address=%D0%BE%D0%BC%D1%81%D0%BA%20%D1%83%D0%BB%20%D0%BA%D1%80%D0%B0%D1%81%D0%BD%D1%8B%D0%B9%20%D0%BF%D1%83%D1%82%D1%8C'

# Получаем HTML-код страницы
page = requests.get(url)

# проверка есть ли доступ к странице
if page.status_code !=200:
   raise Exception("Connection error")

# Создаем объект BeautifulSoup для парсинга HTML-кода
soup = BeautifulSoup(page.text, 'html.parser')

# находим все карточки квартир на странице
cards = soup.findAll("div", class_="flex-item flex-item__properties")

# Создаем новый файл xlsx
workbook = openpyxl.Workbook()
worksheet = workbook.active
worksheet.title = "Квартиры"

# Записываем заголовки в таблицу
worksheet.cell(row=1, column=1, value="Описание")
worksheet.cell(row=1, column=2, value="Площадь")
worksheet.cell(row=1, column=3, value="Цена")

# записываем данные в таблицу
for i, card in enumerate(cards, start=2):
    main_description = card.find("a", {"class":"location slim"}).text.strip()
    price = " ".join(card.find("div", {"class":"property__price"}).text.strip())
    area = card.find("div", {"class":"property__area"}).text.strip()
    
    worksheet.cell(row=i, column=1, value=main_description)
    worksheet.cell(row=i, column=2, value=area)
    worksheet.cell(row=i, column=3, value=price)

# Сохраняем файл
workbook.save("apartments.xlsx")